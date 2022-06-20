#coding:utf-8

import unittest
import os
import openpyxl

class test:
    def __init__(self,file):
        self.file = file

    def test1(self,sheet_name=None):
        if file !=None and sheet_name != None:
            if os.path.isfile(file):
                wb = openpyxl.load_workbook(file)
                # 获取表格所有sheet_name
                sheet = wb[sheet_name]
                return sheet
            elif os.path.isdir(file):
                for maindir, subdir, file_name_list in os.walk(file):
                    for filename in file_name_list:
                        wb = openpyxl.load_workbook(maindir + filename)
                        # 获取表格所有sheet_name
                        sheet_list = wb.sheetnames
                return sheet_list
        elif file != None and sheet_name == None:
            if os.path.isfile(file):
                wb = openpyxl.load_workbook(file)
                # 获取表格所有sheet_name
                sheet_list = wb.sheetnames
                return sheet_list
            elif os.path.isdir(file):
                for maindir, subdir, file_name_list in os.walk(file):
                    for filename in file_name_list:
                        wb = openpyxl.load_workbook(maindir + filename)
                        # 获取所有表格sheet_name
                        sheet_list = wb.sheetnames
                return sheet_list

    def open_excel(self, sheet_name=None):
        """
        打开excel，获取sheet
        :param sheet_name:
        :return:sheet
        """
        wb = openpyxl.load_workbook(self.file)
        if sheet_name !=None:
            # 获取sheet_name
            sheet = wb[sheet_name]
            return sheet
        else:
            sheet1 = wb.sheetnames
            for i in range(len(sheet1)):
                sheet = wb.sheetnames[i]
                sheet = wb[sheet]
                print(sheet)
            return sheet
        # # 获取sheet_name
        # sheet = wb[sheet_name]
        # return sheet

    def get_header(self,sheet_name=None):
        """
        获取header表头
        :param sheet_name:
        :return: header
        """
        if sheet_name !=None:
            wb = self.open_excel(sheet_name)
            header = []
            #遍历第一行
            for i in wb[1]:
                #将遍历出来的表头字段加入列表
                header.append(i.value)
                print(header)
            return header
        else:
            wb = self.open_excel()

    def read_excel(self,sheet_name):
        """
        读取所有数据
        :param sheet_name:
        :return: data
        """
        sheet = self.test1(sheet_name)
        rows = list(sheet.rows)
        data = []
        #遍历从第二行开始的每一行数据
        for row in rows[1:]:
            row_data = []
            #遍历每一行的每个单元格
            for cell in row:
                row_data.append(cell.value)
                #通过zip函数将两个列表合并成字典
                data_dict = dict(zip(self.get_header(sheet_name),row_data))
            data.append(data_dict)
        #返回数据转换为双引号
        #return str(data).replace("\'","\"")
        return data



if __name__ == "__main__":
    file = "D:\Interface_automation\case\logger4.xlsx"
    b = test(file)
    #b.test1()
    #b.open_excel()
    b.get_header()
    print(b)