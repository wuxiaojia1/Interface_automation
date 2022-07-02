#coding:utf-8
import  openpyxl
import json
import ddt
from util.logger_handler import logger
from util.config_handler import yaml_data
import os
class ExcelHandler:

    def __init__(self,file):
        self.file = file

    def open_excel(self,sheet_name):
        """
        打开excel，获取sheet
        :param sheet_name:
        :return:sheet
        """
        wb = openpyxl.load_workbook(self.file)
        #获取sheet_name
        sheet = wb[sheet_name]
        return sheet

    def get_header(self,sheet_name):
        """
        获取header表头
        :param sheet_name:
        :return: header
        """
        wb = self.open_excel(sheet_name)
        header = []
        #遍历第一行
        for i in wb[1]:
            #将遍历出来的表头字段加入列表
            header.append(i.value)
        return header

    def all_sheet_call(self,execl_sheet = None,case_id = None):
        wb = openpyxl.load_workbook(self.file)
        sheets = wb.get_sheet_names()
        data = []
        if execl_sheet == None and case_id == None:
            #循环遍历所有sheet
            for i in sheets:
                #获取每个sheet
                sheet = wb[i]
                rows =list(sheet.rows)
                # 遍历从第二行开始的每一行数据
                for row in rows[1:]:
                    row_data = []
                    # 遍历每一行的每个单元格
                    for cell in row:
                        row_data.append(cell.value)
                        # 通过zip函数将两个列表合并成字典
                        data_dict = dict(zip(self.get_header(i), row_data))
                    data.append(data_dict)
            return data
        elif execl_sheet != None and case_id == None:
            sheet_name = wb[execl_sheet]
            rows = list(sheet_name.rows)
            #遍历从第二行开始的每一行数据
            for row in rows[1:]:
                row_data = []
                #遍历每一行的每个单元格
                for cell in row:
                    row_data.append(cell.value)
                    #通过zip的函数将两个列表合并成字典
                    data_dict = dict(zip(self.get_header(execl_sheet),row_data))
                data.append(data_dict)
            return data
        elif execl_sheet == None and case_id != None:
            for i in sheets:
                #获取每个sheet
                sheet = wb[i]
                rows =list(sheet.rows)
                # 遍历从第二行开始的每一行数据
                for row in rows[1:]:
                    row_data = []
                    ## 遍历每一行的每个单元格
                    for cell in row:
                        row_data.append(cell.value)
                        data_dict = dict(zip(self.get_header(i),row_data))
                    if data_dict["case_id"] == case_id:
                        data.append(data_dict)
                return data
        else:
            sheet_name = wb[execl_sheet]
            rows = list(sheet_name.rows)
            for row in rows[1:]:
                row_data = []
                for cell in row:
                    row_data.append(cell.value)
                    data_dict = dict(zip(self.get_header(execl_sheet),row_data))
                if data_dict['case_id'] == case_id:
                    data.append(data_dict)
            return data






    def read_excel(self,sheet_name):
        """
        读取所有数据
        :param sheet_name:
        :return: data
        """
        sheet = self.open_excel(sheet_name)
        rows = list(sheet.rows)
        data = []
        #遍历从第二行开始的每一行数据
        for row in rows[1:]:
            row_data = []
            #遍历每一行的每个单元格
            for cell in row:
                row_data.append(cell.value)
                #通过zip函数将两个列表合并成字典
                data_dict = dict(zip(self.get_header(sheet_name),row_data))
            data.append(data_dict)
        #返回数据转换为双引号
        #return str(data).replace("\'","\"")
        return data

    @staticmethod
    def write_excel(file,sheet_name,row,cloumn,data):
        """
        excel写入数据
        :param file:
        :param sheet_name:
        :param row:
        :param cloumn:
        :param data:
        :return:
        """
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet_name]
        sheet.cell(row,cloumn).value = data
        wb.save(file)
        wb.close()

if __name__ == '__main__':
    file = yaml_data['case']['file']
    sheet = yaml_data['case']['sheet']
    case_id = yaml_data['case']['case_id']
    excel = ExcelHandler(file)
    print(yaml_data['case']['file'])
    #data = excel.open_excel('login')
    #data1 = excel.open_excel("login")
    #data2 = excel.read_excel("login")
    #print(data1)
    # data1 = excel.get_header('login')
    # print(data)
    # logger.info("当前是第{0}条用例:{1}".format('nihao','henhao'))
    a = excel.all_sheet_call(case_id=case_id)
    print(a)

    # write = excel.write_excel("D:\Interface_automation\case\logger4.xlsx", 'login',2,9, 'haha')




