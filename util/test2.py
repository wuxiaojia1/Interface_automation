#coding:utf-8
import os
import openpyxl

file = 'D:\Interface_automation\case\logger4.xlsx'

wb = openpyxl.load_workbook(file)
a = wb.sheetnames
print(a)
print(type(a))
for i in range(len(a)):
    print(i)
    b = wb.sheetnames[i]
    c = wb[b]
    print(c)